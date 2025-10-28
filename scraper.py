import requests
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET
import json
import os
import re
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
import tweepy
from dotenv import load_dotenv
import logging
from urllib.parse import urlparse
from datetime import datetime
from openpyxl import Workbook, load_workbook
import sys
from instagrapi import Client
import cv2
import csv
import argparse
from typing import Dict, List

load_dotenv()

DATA = os.getenv("DATA")
HASHTAG = os.getenv("HASHTAG")
LOGO = os.getenv("LOGO")
RSS_URL = os.getenv("RSS_URL")
FONT_1 = os.getenv("FONT_1")
FONTSIZE_1 = os.getenv("FONTSIZE_1")
FONTSIZE_1 = int(os.getenv("FONTSIZE_1"))
FONT_2 = os.getenv("FONT_2")
FONTSIZE_2 = int(os.getenv("FONTSIZE_2"))
FONT_3 = os.getenv("FONT_3")
FONTSIZE_3 = int(os.getenv("FONTSIZE_3"))
LOG = os.getenv("LOG")
IG_FIELDS = ("ig_username", "ig_password")
FB_FIELDS = ("fb_id", "fb_token")
X_FIELDS = ("x_key", "x_keysecret", "x_access", "x_accesstoken", "x_bearertoken")

logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
file_handler = logging.FileHandler(LOG)
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

def fetch_rss(url):
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})
    if response.status_code == 200:
        return response.content
    else:
        raise Exception(f"Failed to fetch RSS: {response.status_code}")

def parse_news_urls(rss_content):
    root = ET.fromstring(rss_content)
    ns = {'ht': 'https://trends.google.com/trending/rss'}
    urls = []
    for item in root.findall('./channel/item'):
        for news_item in item.findall('ht:news_item', ns):
            url_tag = news_item.find('ht:news_item_url', ns)
            if url_tag is not None and url_tag.text:
                urls.append(url_tag.text)
    return urls


# === Fungsi pembersih teks ===
def clean_text(text):
    if not text:
        return ""

    text = text.replace('\u2013', '-')   # en dash
    text = text.replace('\u2014', '-')   # em dash
    text = text.replace('\u00A0', ' ')   # non-breaking space
    text = text.replace('\u200B', '')    # zero-width space
    text = text.replace('\u2026', '...') # ellipsis

    text = re.sub(r'[^\x20-\x7E\u00A0-\u00FF\u0100-\u017F\u0180-\u024F\u1E00-\u1EFF]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


# === 2. Scraper untuk setiap link ===
def scrape_website(url):
    try:
        response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        if response.status_code != 200:
            logging.error(f"[x] Gagal ambil {url}: {response.status_code}")
            return None

        soup = BeautifulSoup(response.text, "html.parser")

        title = clean_text(soup.title.string.strip()) if soup.title and soup.title.string else None

        paragraphs = []
        for p in soup.find_all("p"):
            # text = clean_text(p.get_text(strip=True))
            text = clean_text(p.get_text(strip=False))
            if len(text) > 50:
                paragraphs.append(text)

        meta_data = {}
        for prop in ["og:title", "og:image", "og:description"]:
            tag = soup.find("meta", property=prop)
            meta_data[prop] = clean_text(tag["content"]) if tag and tag.has_attr("content") else None

        return {
            "link": url,
            "title": title,
            "paragraphs": paragraphs,
            "meta": meta_data
        }

    except Exception as e:
        logging.error(f"[!] Error scraping {url}: {e}")
        return None


# === 3. Simpan ke JSON tanpa duplikasi ===
def save_to_json(data, filename="scraped_result.json"):
    if not data:
        return False

    data["status"] = "pending"  # Tambahkan status awal
    data["date"] = datetime.now().strftime("%Y-%m-%d %I:%M:%S")

    if os.path.exists(filename):
        with open(filename, "r", encoding="utf-8") as f:
            try:
                all_data = json.load(f)
            except json.JSONDecodeError:
                all_data = []
    else:
        all_data = []

    if not any(entry["link"] == data["link"] for entry in all_data):
        all_data.append(data)
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(all_data, f, indent=4, ensure_ascii=False)

        return True
    else:
        logging.info("skip.")
        return False

def update_status_json(link, status="done", post_id=None, filename="scraped_result.json"):
    """Update status pada JSON berdasarkan link."""
    if not os.path.exists(filename):
        logging.error("[!] File JSON tidak ditemukan.")
        return

    try:
        with open(filename, "r", encoding="utf-8") as f:
            all_data = json.load(f)

        updated = False
        for entry in all_data:
            if entry.get("link") == link:
                entry["status"] = status
                if post_id:
                    entry["post_id"] = post_id
                updated = True
                break

        if updated:
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(all_data, f, indent=4, ensure_ascii=False)
            logging.info(f"[✓] Status '{status}' diupdate untuk: {link}")
        else:
            logging.warning(f"[!] Link {link} tidak ditemukan di JSON.")

    except Exception as e:
        logging.error(f"[!] Gagal update status JSON: {e}")

# === 4. Fungsi bantu untuk download gambar OG:image ===
def get_background_image(og_image_url, default_path="background.png"):
    if not og_image_url:
        return default_path

    try:
        response = requests.get(og_image_url, timeout=10)
        if response.status_code == 200:
            img = Image.open(BytesIO(response.content))
            if img.mode == "RGBA":
                img = img.convert("RGB")  # ubah ke RGB agar bisa disimpan ke JPEG

            temp_path = get_safe_filename_from_url(og_image_url) + ".jpg"
            img.save(temp_path, "JPEG")
            return temp_path
    except Exception as e:
        logging.error(f"[!] Gagal ambil gambar background: {e}")
    return default_path


# === 5. Fungsi pembuat poster ===
def wrap_text(draw, text, font, max_width):
    words = text.split()
    lines, line = [], ""
    for word in words:
        test_line = line + " " + word if line else word
        bbox = draw.textbbox((0, 0), test_line, font=font)
        w = bbox[2] - bbox[0]
        if w <= max_width:
            line = test_line
        else:
            lines.append(line)
            line = word
    if line:
        lines.append(line)
    return lines


def buat_poster(content_title, content_body, hashtag, logo_path, bg_path, output_file):

    try:
        bg = Image.open(bg_path).convert("RGB")
    except Exception as e:
        logging.error(f"[x] Gagal load background: {e}")
        return

    width, height = bg.size
    font_hashtag = ImageFont.truetype(FONT_1, FONTSIZE_1)
    font_title = ImageFont.truetype(FONT_2, FONTSIZE_2)
    font_body = ImageFont.truetype(FONT_3, FONTSIZE_3)

    dummy_img = Image.new("RGB", (width, height))
    dummy_draw = ImageDraw.Draw(dummy_img)

    lines_title = wrap_text(dummy_draw, content_title, font_title, int(width*0.9))
    lines_body = wrap_text(dummy_draw, content_body, font_body, int(width*0.9))

    spacing_title, spacing_body = 5, 6
    total_text_height = (len(lines_title) * (font_title.size + spacing_title) +
                         len(lines_body) * (font_body.size + spacing_body))

    padding_top_bottom = 40
    footer_height = total_text_height + padding_top_bottom*2 + 80

    new_height = height + footer_height
    img = Image.new("RGB", (width, new_height), "white")
    img.paste(bg, (0, 0))

    draw = ImageDraw.Draw(img)
    footer_y = height
    draw.rectangle([0, footer_y, width, new_height], fill=(253, 34, 66))

    # Logo
    try:
        logo = Image.open(logo_path).convert("RGBA")
        max_logo_width = int(width * 0.15)
        ratio = max_logo_width / logo.width
        logo = logo.resize((max_logo_width, int(logo.height * ratio)))
        img.paste(logo, (30, 30), logo)
    except Exception as e:
        logging.error("Gagal load logo:", e)

    bbox = draw.textbbox((0, 0), hashtag, font=font_hashtag)
    text_w, text_h = bbox[2] - bbox[0], bbox[3] - bbox[1]

    pad_x, pad_y = 20, 10
    box_x1 = (width - text_w)//2 - pad_x
    box_y1 = footer_y + 20
    box_x2 = box_x1 + text_w + pad_x*2
    box_y2 = box_y1 + text_h + pad_y*2

    draw.rectangle([box_x1, box_y1, box_x2, box_y2], fill="white")
    draw.text(((width-text_w)//2, box_y1+pad_y), hashtag,
              font=font_hashtag, fill=(253,34,66))

    y_text = box_y2 + 30
    for line in lines_title:
        draw.text((50, y_text), line, font=font_title, fill="white")
        y_text += font_title.size + spacing_title

    y_text += 20
    for line in lines_body:
        draw.text((50, y_text), line, font=font_body, fill="white")
        y_text += font_body.size + spacing_body

    img.save(output_file)

def tweet_image(x_key, x_keysecret, x_access, x_accesstoken, x_bearertoken, image_path, tweet_text):
    """
    Upload gambar dan buat tweet di X (Twitter).

    Args:
        image_path (str): Path file gambar yang akan diunggah.
        tweet_text (str): Isi tweet.
    """

    try:
        # --- Autentikasi OAuth 1.0a untuk upload media ---
        auth = tweepy.OAuth1UserHandler(
            x_key, x_keysecret,
            x_access, x_accesstoken
        )
        api = tweepy.API(auth)

        # --- Upload media ---
        media = api.media_upload(image_path)

        # --- Client v2 untuk bikin tweet ---
        client = tweepy.Client(
            bearer_token=x_bearertoken,
            consumer_key=x_key,
            consumer_secret=x_keysecret,
            access_token=x_access,
            access_token_secret=x_accesstoken
        )

        # --- Post tweet dengan foto ---
        response = client.create_tweet(
            text=tweet_text,
            media_ids=[media.media_id]
        )

        tweet_id = response.data["id"]
        tweet_url = f"https://x.com/user/status/{tweet_id}"
        logging.info(tweet_url)

        return tweet_id

    except Exception as e:
        logging.error(f"[!] Gagal upload tweet: {e}")

def upload_photo_facebook(page_id, page_token, file_path, caption=""):
    """
    Upload foto lokal ke Facebook Page menggunakan Graph API.
    
    Args:
        page_id (str): ID halaman Facebook
        page_token (str): Access token halaman
        file_path (str): Path file gambar lokal
        caption (str): Caption untuk foto
    
    Returns:
        dict: Hasil respon JSON dari Facebook API
    """
    url = f"https://graph.facebook.com/v20.0/{page_id}/photos"
    payload = {
        "caption": caption or "Upload foto 🚀",
        "access_token": page_token
    }

    try:
        with open(file_path, "rb") as f:
            files = {"source": f}
            res = requests.post(url, data=payload, files=files)

        # Coba ambil JSON-nya
        try:
            response_data = res.json()
        except Exception:
            response_data = {"error": "Tidak bisa parse respon", "raw": res.text}

        # Cek status kode
        if res.status_code == 200:
            print(f"✅ Berhasil upload ke Facebook Page ID {page_id}")
        else:
            print(f"❌ Gagal upload (status {res.status_code})")

        return response_data

    except FileNotFoundError:
        print(f"⚠️ File tidak ditemukan: {file_path}")
        return {"error": "File tidak ditemukan"}

    except Exception as e:
        print(f"⚠️ Terjadi kesalahan: {e}")
        return {"error": str(e)}

# def get_safe_filename_from_url(url):
#     parsed = urlparse(url)
#     basename = os.path.basename(parsed.path)
#     name, _ = os.path.splitext(basename)
#     safe_name = re.sub(r'[<>:"/\\|?*]', '_', name)
#     safe_name = safe_name.strip().replace(' ', '_')
#     return safe_name

def get_safe_filename_from_url(url):
    parsed = urlparse(url)
    path = parsed.path.rstrip("/")  # hapus slash di akhir jika ada
    basename = os.path.basename(path)
    name, _ = os.path.splitext(basename)

    # Kalau kosong, fallback pakai domain (biar gak kosong total)
    if not name:
        name = parsed.netloc.replace('.', '_')

    safe_name = re.sub(r'[<>:"/\\|?*]', '_', name)
    safe_name = safe_name.strip().replace(' ', '_')
    return safe_name

def save_to_excel(data, filename="data_sosmed.xlsx"):
    """
    Menyimpan data ke file Excel dengan header:
    no | sosmed | username | url site | url sosmed | date time post | status
    
    Jika file belum ada, otomatis dibuat dengan header.
    Kolom 'no' akan auto increment berdasarkan data terakhir.
    """

    headers = ["no", "sosmed", "username", "url site", "url sosmed", "date time post", "status"]

    # Jika file belum ada, buat workbook baru
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(filename)

    # Load file Excel yang sudah ada
    wb = load_workbook(filename)
    ws = wb.active

    # Ambil nomor terakhir di kolom pertama (no)
    last_row = ws.max_row
    if last_row > 1:
        last_no = ws.cell(row=last_row, column=1).value
        next_no = last_no + 1 if isinstance(last_no, int) else 1
    else:
        next_no = 1

    # Isi data baru
    row_data = [
        next_no,
        data.get("sosmed", ""),
        data.get("username", ""),
        data.get("url_site", ""),
        data.get("url_sosmed", ""),
        data.get("date_time_post", datetime.now().strftime("%Y-%m-%d %I:%M:%S")),
        data.get("status", "done")
    ]

    ws.append(row_data)
    wb.save(filename)

    log = f"[+] Data berhasil disimpan ke {filename} (no={next_no})"
    logging.info(log)

# === 2. Upload media (foto/video) ===
def upload_media(username: str, password: str = None, image_path: str = None, video_path: str = None, caption: str = ""):
    cl = Client()

    # Selalu login ulang (tidak pakai load_session)
    print("[*] Login ke akun Instagram...")
    try:
        cl.login(username, password)
        print("[+] Login berhasil!")
    except Exception as e:
        print("[!] Gagal login:", e)
        return False

    # === Upload ===
    try:
        if image_path:
            if not os.path.exists(image_path):
                print(f"[!] File gambar tidak ditemukan: {image_path}")
                return False
            print("[*] Upload foto...")
            media = cl.photo_upload(image_path, caption)

        elif video_path:
            if not os.path.exists(video_path):
                print(f"[!] File video tidak ditemukan: {video_path}")
                return False

            # Buat thumbnail otomatis
            thumb_path = os.path.splitext(video_path)[0] + "_thumb.jpg"
            if not os.path.exists(thumb_path):
                print("[*] Membuat thumbnail otomatis...")
                cap = cv2.VideoCapture(video_path)
                ret, frame = cap.read()
                if ret:
                    cv2.imwrite(thumb_path, frame)
                cap.release()

            print("[*] Upload video...")
            media = cl.video_upload(video_path, caption, thumbnail=thumb_path)
        else:
            print("[!] Harus isi image_path atau video_path.")
            return False

        return getattr(media, "code", None)

    except Exception as e:
        print("[!] Upload gagal:", e)
        return False

    finally:
        try:
            cl.logout()
        except Exception:
            pass

#!/usr/bin/env python3
"""
extract_social_sources.py
Membaca CSV baris-per-baris. Untuk setiap baris:
 - Deteksi source(s) berdasarkan field yang berisi data
 - Jika IG terdeteksi (ig_username & ig_password ada) -> source = "IG"
 - Jika FB terdeteksi (fb_id & fb_token ada) -> source = "FB"
 - Jika X/Twitter terdeteksi (ada kombinasi kunci/tokens) -> source = "X"
 - Kembalikan hasil ke sebuah JSON/CSV output

Contoh pemakaian:
    python extract_social_sources.py input.csv output.json

File ini tidak benar-benar melakukan login ke platform apa pun — hanya mengekstrak
kredensial / identifier dan menandai sumbernya. Anda bisa menambahkan fungsi
fetch_from_ig / fetch_from_fb / fetch_from_x sesuai kebutuhan (API calls, SDK, dsb).
"""


def has_values(row: Dict[str, str], fields) -> bool:
    """Return True if ALL fields exist in row and are non-empty after strip()."""
    for f in fields:
        val = row.get(f, "")
        if val is None or str(val).strip() == "":
            return False
    return True


def detect_sources(row: Dict[str, str]) -> List[str]:
    """Detect which social sources are present for a given CSV row.

    Returns a list of source codes: e.g. ["IG"], ["FB"], ["X"], or a combination.
    """
    sources = []
    if has_values(row, IG_FIELDS):
        sources.append("IG")
    if has_values(row, FB_FIELDS):
        sources.append("FB")
    # For X we accept partial credentials as long as at least 1 token/key present
    # but you can change the logic to require ALL fields.
    if any(str(row.get(f, "")).strip() != "" for f in X_FIELDS):
        sources.append("X")
    return sources


def build_payload_for_source(row: Dict[str, str], source: str) -> Dict[str, str]:
    """Return a cleaned dict of the relevant fields for a source."""
        
    return {
        "ig_username": row.get("ig_username", "").strip(),
        "ig_password": row.get("ig_password", "").strip(),
        "fb_id": row.get("fb_id", "").strip(),
        "fb_token": row.get("fb_token", "").strip(),
    }
    
    if source == "FB":
        return {
            "fb_id": row.get("fb_id", "").strip(),
            "fb_token": row.get("fb_token", "").strip(),
        }
    if source == "X":
        return {f: row.get(f, "").strip() for f in X_FIELDS}
    # fallback: return empty
    return {}

def process_row(row: Dict[str, str]) -> Dict:
    """Detect sources, build payloads and optionally call fetch functions.

    Returned dict example:
    {
      "username": "Breaking News",
      "sources": ["IG"],
      "credentials": {"IG": {...}},
      "fetch_results": {"IG": {...}}  # optional
    }
    """
    username = row.get("username", "").strip()
    detected = detect_sources(row)
    result = dict(row)

    return {
        "username": username,
        "sources": detected,
        "result": result,
    }

def read_csv(path: str):
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = []
        for r in reader:
            clean = {}
            for k, v in r.items():
                if k is None:
                    continue  # lewati kolom tanpa nama
                clean[k.strip()] = (v or "").strip()
            rows.append(clean)
    return rows

def save_json(data, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

import time

# === 6. Jalankan semua ===
if __name__ == "__main__":

    logging.info("Fetch RSS")
    rss_content = fetch_rss(RSS_URL)
    news_urls = parse_news_urls(rss_content)
    logging.info(f"{len(news_urls)} data")

    for i, url in enumerate(news_urls, start=1):
        logging.info(f"[{i}] {url}")
        result = scrape_website(url)
        if result and save_to_json(result):
            title = result["title"] or "Berita Trending"

            body = " ".join(result["paragraphs"][:2]) if result["paragraphs"] else (result["meta"].get("og:description") or "")
            body = body.strip()

            if len(body) < 300:
                clean_body = body
            else:
                truncated = body[:300]
                if not truncated.endswith(('.', '!', '?')):
                    last_period = truncated.rfind('.')
                    if last_period != -1:
                        clean_body = truncated[:last_period+1].strip()
                    else:
                        clean_body = truncated.strip()
                else:
                    clean_body = truncated.strip()

            if len(clean_body) < 300 and not clean_body.endswith(('.', '!', '?')):
                sentences = re.split(r'(?<=[.!?])\s+', body)
                if len(sentences) > 1:
                    clean_body = sentences[-2].strip() + " " + sentences[-1].strip()

            bg_path = get_background_image(result["meta"].get("og:image"))
            output_name = get_safe_filename_from_url(url) + ".jpg"

            logging.info("Start IMG")
            buat_poster(
                title,
                clean_body,
                hashtag=HASHTAG,
                logo_path=LOGO,
                bg_path=bg_path,
                output_file=output_name
            )
            logging.info("Success IMG")

            logging.info("Start Post")
            user_name = [s.strip() for s in USER_NAME.split(",") if s.strip()]
            social_media = [s.strip() for s in SOCIAL_MEDIA.split(",") if s.strip()]
            ig_username = [s.strip() for s in IG_USERNAME.split(",") if s.strip()]
            ig_password = [s.strip() for s in IG_PASSWORD.split(",") if s.strip()]

            rows = read_csv(DATA)
            results = []
            for i, r in enumerate(rows, start=1):
                res = process_row(r)
                url_site = result["link"]
                username = res["username"]
                sources = res["sources"][0]
                credentials = res["result"]
                ig_username = credentials["ig_username"]
                ig_password = credentials["ig_password"]
                fb_id = credentials["fb_id"]
                fb_token = credentials["fb_token"]
                x_key = credentials["x_key"]
                x_keysecret = credentials["x_keysecret"]
                x_access = credentials["x_access"]
                x_accesstoken = credentials["x_accesstoken"]
                x_bearertoken = credentials["x_bearertoken"]
                status = "pending"
                post_id = ""
                url_sosmed = ""

                if sources == "IG":
                    post_id = upload_media(ig_username, ig_password, output_name, title)
                    url_sosmed = "https://instagram.com/p/" + post_id

                if sources == "X":
                    post_id = tweet_image(x_key, x_keysecret, x_access, x_accesstoken, x_bearertoken, output_name, title)
                    url_sosmed = "https://x.com/user/status/" + post_id
                
                if sources == "FB":
                    post_id = upload_photo_facebook(fb_id, fb_token, output_name, title)
                    post_id = post_id.get("id")
                    url_sosmed = "https://fb.com/" + post_id
         
                if post_id != "":
                    status = "done"
                    update_status_json(
                        url_site, 
                        status=status,
                        post_id=post_id
                    )

                logging.info("Success Post")

                new_data = {
                    "sosmed": sources,
                    "username": username,
                    "url_site": url_site,
                    "url_sosmed": url_sosmed,
                    "status": status
                }

                save_to_excel(new_data)
                time.sleep(100000)
                
